from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse, FileResponse
import pandas as pd
import os
from typing import Dict, List, Optional, Tuple
from pydantic import BaseModel
import uvicorn
import chardet
import logging
from openpyxl import load_workbook
from datetime import datetime, timedelta
import hashlib

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="API de Arquivos com Atualização Automática")

# Configurações
BASE_PATHS = [
    r"G:\Meu Drive\Arqs.Dia",
    r"H:\Meu Drive\Arqs.Dia"
]
CACHE_TIME = timedelta(minutes=5) 


# ESTRUTURA DE CACHE INTELIGENTE
class FileCache:
    def __init__(self):
        self.data = {}
        self.last_checked = {}
        self.file_hashes = {}

    def needs_refresh(self, filepath: str) -> bool:
        """Verifica se o arquivo precisa ser atualizado"""
        if filepath not in self.last_checked:
            return True
            
        if datetime.now() - self.last_checked[filepath] > CACHE_TIME:
            try:
                current_hash = self._calculate_file_hash(filepath)
                return current_hash != self.file_hashes.get(filepath)
            except Exception as e:
                logger.warning(f"Erro ao verificar hash: {str(e)}")
                return True
        return False

    def _calculate_file_hash(self, filepath: str) -> str:
        """Calcula hash MD5 do arquivo"""
        hash_md5 = hashlib.md5()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def update_cache(self, filepath: str, data: dict):
        """Atualiza o cache"""
        self.data[filepath] = data
        self.last_checked[filepath] = datetime.now()
        self.file_hashes[filepath] = self._calculate_file_hash(filepath)

cache = FileCache()

# CONFIGURAÇÃO DOS ARQUIVOS
FILE_CONFIG = {
    "Base_GDM": {"filename": "Base_GDM.xlsx", "type": "excel"},
    "BASE_Grupo_VD": {"filename": "BASE_Grupo_VD.xlsx", "type": "excel"},
    "Base_IAV": {"filename": "Base_IAV.xlsx", "type": "excel"},
    "Base_ID": {"filename": "Base_ID.xlsx", "type": "excel", "required_sheet": "RESUMO"},
    "Base_INE": {"filename": "Base_INE.xlsx", "type": "excel", "required_sheet": "RESUMO"},
    "BASE_MOV": {"filename": "BASE_MOV.xlsx", "type": "excel"},
    "Base_MOV_AA": {"filename": "Base_MOV_AA.xlsx", "type": "excel"},
    "Base_MOV_ANT": {"filename": "Base_MOV_ANT.xlsx", "type": "excel"},
    "BASE_PDV": {"filename": "BASE_PDV.xlsx", "type": "excel"},
    "Base_PROD": {"filename": "Base_PROD.xlsx", "type": "excel"},
    "Base_TEND": {"filename": "Base_TEND.xlsx", "type": "excel"},
    "PRODUTOS": {
        "filename": "PRODUTOS.csv", 
        "type": "csv", 
        "encoding": "ISO-8859-1",
        "delimiter": ";",
        "on_bad_lines": "skip"
    },
    "BASE_MKP_VD": {
        "filename": "BASE_MKP_VD.txt", 
        "type": "txt", 
        "delimiter": "\t",
        "encoding": "utf-8"
    },
    "BASE_MKP_VD_AA": {
        "filename": "BASE_MKP_VD_AA.txt", 
        "type": "txt", 
        "delimiter": "|",
        "encoding": "ISO-8859-1",
        "on_bad_lines": "skip"
    }
}

# FUNÇÕES AUXILIARES
def find_file(filename: str) -> Optional[str]:
    """Localiza arquivos nos diretórios configurados"""
    for path in BASE_PATHS:
        filepath = os.path.join(path, filename)
        if os.path.exists(filepath):
            return filepath
    return None

def safe_read_file(filepath: str, config: dict) -> pd.DataFrame:
    """Lê arquivos com tratamento robusto de erros"""
    try:
        if config["type"] == "excel":
            if "required_sheet" in config:
                wb = load_workbook(filepath, read_only=True)
                if config["required_sheet"] not in wb.sheetnames:
                    available = ", ".join(wb.sheetnames)
                    raise ValueError(f"Guia '{config['required_sheet']}' não encontrada. Disponíveis: {available}")
                return pd.read_excel(filepath, sheet_name=config["required_sheet"], engine='openpyxl')
            return pd.read_excel(filepath, engine='openpyxl')
        
        elif config["type"] == "csv":
            return pd.read_csv(
                filepath,
                delimiter=config.get("delimiter", ";"),
                encoding=config.get("encoding", "utf-8"),
                on_bad_lines=config.get("on_bad_lines", "warn")
            )
        else:  # TXT
            return pd.read_csv(
                filepath,
                delimiter=config.get("delimiter", "\t"),
                encoding=config.get("encoding", "utf-8"),
                on_bad_lines=config.get("on_bad_lines", "warn"),
                quoting=3
            )
    except Exception as e:
        logger.error(f"Falha na leitura de {filepath}: {str(e)}")
        raise

# ENDPOINTS PRINCIPAIS
@app.on_event("startup")
async def startup_event():
    """Carrega todos os arquivos ao iniciar"""
    logger.info("Iniciando carga inicial...")
    for file_id in FILE_CONFIG.keys():
        try:
            await _load_file_data(file_id)
        except Exception as e:
            logger.error(f"Erro ao carregar {file_id}: {str(e)}")

async def _load_file_data(file_id: str) -> Tuple[bool, dict]:
    """Carrega/atualiza dados de um arquivo"""
    config = FILE_CONFIG[file_id]
    filepath = find_file(config["filename"])
    if not filepath:
        raise HTTPException(404, detail="Arquivo não encontrado")

    if not cache.needs_refresh(filepath):
        return False, cache.data[filepath]

    try:
        df = safe_read_file(filepath, config)
        df = df.replace([pd.NA, pd.NaT, float('nan'), float('inf')], None)
        
        for col in df.select_dtypes(include=['datetime']).columns:
            df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')

        data = {
            "data": df.to_dict(orient="records"),
            "metadata": {
                "last_updated": datetime.now().isoformat(),
                "file_size": f"{os.path.getsize(filepath)/1024/1024:.2f} MB",
                "row_count": len(df)
            }
        }
        
        cache.update_cache(filepath, data)
        return True, data
    except Exception as e:
        logger.error(f"Erro ao processar {file_id}: {str(e)}")
        raise

@app.get("/refresh/{file_id}")
async def refresh_file(file_id: str, background_tasks: BackgroundTasks):
    """Força atualização de um arquivo específico"""
    background_tasks.add_task(_load_file_data, file_id)
    return {"message": f"Atualização de {file_id} em andamento"}

@app.get("/refresh-all")
async def refresh_all(background_tasks: BackgroundTasks):
    """Força atualização de todos os arquivos"""
    for file_id in FILE_CONFIG.keys():
        background_tasks.add_task(_load_file_data, file_id)
    return {"message": "Atualização completa em andamento"}

@app.get("/data/{file_id}")
async def get_file_data(
    file_id: str,
    skiprows: int = 0,
    nrows: Optional[int] = None,
    as_excel: bool = False,
    background_tasks: BackgroundTasks = None
):
    """Obtém dados com atualização automática"""
    if background_tasks:
        background_tasks.add_task(_load_file_data, file_id)
    
    config = FILE_CONFIG[file_id]
    filepath = find_file(config["filename"])
    if not filepath:
        raise HTTPException(404, detail="Arquivo não encontrado")

    try:
        refreshed, data = await _load_file_data(file_id)
        df = pd.DataFrame(data["data"])
        
        if skiprows > 0:
            df = df.iloc[skiprows:]
        if nrows is not None:
            df = df.head(nrows)

        if as_excel:
            output = f"temp_{file_id}.xlsx"
            df.to_excel(output, index=False)
            return FileResponse(output, filename=f"{file_id}.xlsx")
            
        return JSONResponse({
            "data": df.to_dict(orient="records"),
            "metadata": data["metadata"]
        })
    except Exception as e:
        raise HTTPException(500, detail=f"Erro: {str(e)}")

@app.get("/files")
async def list_files():
    """Lista todos os arquivos disponíveis"""
    available = []
    for file_id, config in FILE_CONFIG.items():
        path = find_file(config["filename"])
        if path:
            stat = os.stat(path)
            available.append({
                "name": file_id,
                "type": config["type"],
                "size": f"{stat.st_size/1024/1024:.2f} MB",
                "last_modified": datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M'),
                "path": path
            })
    return available

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
